from openpyxl import load_workbook

wb = load_workbook("Family_lists.xlsx")
sheet = wb.active

# i=0
# for row in sheet.iter_rows(min_row=2, max_row=2, min_col = 1, max_col=5):
#     for cell in row:
#         cell.value = i
#         i += 1

# wb.remove_sheet(sheet)
# print(sheet.title)

sheet.delete_rows(1)

# wb.copy_worksheet(sheet)

# for row in sheet.iter_rows():
#     for cell in row:
#         print(cell.value, end = " ")
#     print()
wb.save("Family_lists.xlsx")
wb.close()


# ## ==> MESSBOX WARNING
# from tkinter import messagebox, Tk
# Tk().withdraw()

# messagebox.showwarning("Увага!", "Інформація не введена в одне з обов'язкових полів!")
