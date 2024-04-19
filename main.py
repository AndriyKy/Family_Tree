import sys

from openpyxl import Workbook, load_workbook
from PyQt5 import QtWidgets

from business_logic import (
    AddEdit,
    AddRemoveClan,
    FamilyTies,
    Graph,
    MyCard,
    Review,
)
from business_logic.constructor import WORKBOOK_NAME
from pyui import UIMainWindow


def open() -> None:
    app = QtWidgets.QApplication(sys.argv)

    main_window = QtWidgets.QMainWindow()
    ui_main_window = UIMainWindow()
    ui_main_window.setupUi(main_window)
    main_window.show()

    my_card = MyCard(ui_main_window.pushButton_myCard)
    add_remove_clan = AddRemoveClan()
    add_edit = AddEdit()
    family_ties = FamilyTies()
    review = Review()
    graph = Graph()

    # Register window opening events.
    ui_main_window.pushButton_myCard.clicked.connect(my_card.openEvent)
    ui_main_window.pushButton_addRemoveClan.clicked.connect(
        add_remove_clan.openEvent
    )
    ui_main_window.pushButton_addEdit.clicked.connect(add_edit.openEvent)
    ui_main_window.pushButton_familyTies.clicked.connect(family_ties.openEvent)
    ui_main_window.pushButton_review.clicked.connect(review.openEvent)
    ui_main_window.pushButton_graph.clicked.connect(graph.openEvent)
    # Condition of pressing buttons / changing combo boxes.
    my_card.ui_window.image_Button.clicked.connect(my_card.set_avatar)
    my_card.ui_window.pushButton_done.clicked.connect(my_card.save)
    add_remove_clan.ui_window.pushButton_done.clicked.connect(
        add_remove_clan.save
    )

    add_edit.ui_window.comboBox_addEdit.currentIndexChanged.connect(
        add_edit.refill_window
    )
    add_edit.ui_window.image_Button.clicked.connect(add_edit.set_avatar)
    add_edit.ui_window.pushButton_done.clicked.connect(add_edit.save)
    family_ties.ui_window.pushButton_addSiblings.clicked.connect(
        family_ties.add_sibling_to_table
    )
    family_ties.ui_window.pushButton_done.clicked.connect(family_ties.save)
    review.ui_window.pushButton_done.clicked.connect(review.close)

    try:
        wb = load_workbook(WORKBOOK_NAME)
        my_card.worksheet = wb.active
        my_card.set_my_card_button_label()
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.insert_rows(0)
        wb.save(WORKBOOK_NAME)
        my_card.openEvent()
    finally:
        wb.close()

    app.exec()


if __name__ == "__main__":
    open()
